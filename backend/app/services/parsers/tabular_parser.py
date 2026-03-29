import csv
from pathlib import Path

from openpyxl import load_workbook

from app.services.parsers.base import BaseParser, ParsedDocument


class TabularParser(BaseParser):
    supported_extensions = {".csv", ".xlsx"}

    @staticmethod
    def supported_mime_types() -> set[str]:
        return {
            "text/csv",
            "application/csv",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }

    def parse(self, path: Path) -> ParsedDocument:
        extension = path.suffix.lower()
        if extension == ".csv":
            return self._parse_csv(path)
        if extension == ".xlsx":
            return self._parse_xlsx(path)

        raise ValueError(f"Unsupported tabular extension: {extension}")

    def _parse_csv(self, path: Path) -> ParsedDocument:
        rows: list[str] = []
        with path.open("r", encoding="utf-8", errors="ignore") as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                rows.append(" | ".join(cell.strip() for cell in row))

        return ParsedDocument(text="\n".join(rows), page_count=1)

    def _parse_xlsx(self, path: Path) -> ParsedDocument:
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        row_count = 0

        for sheet in workbook.worksheets:
            lines.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_values = ["" if value is None else str(value) for value in row]
                lines.append(" | ".join(row_values).rstrip(" |"))
                row_count += 1

        workbook.close()
        return ParsedDocument(text="\n".join(lines), page_count=max(1, row_count))
